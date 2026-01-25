# utils/loaders/xlsx_loader.py

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Mapping

from openpyxl import load_workbook

from .base_loader import (
    BaseVendorLoader,
    LoaderReadError,
    LoaderMetadataError,
    LoaderNormalizationError,
)


class XLSXLoader(BaseVendorLoader):
    """
    Universal XLSX loader.

    Loads Excel workbooks into a dict:
        {
            "Sheet1": [ {col: value}, ... ],
            "Sheet2": [ ... ],
        }
    """

    VENDOR = "universal"
    FORMAT = "xlsx"
    EXTENSIONS = (".xlsx", ".xlsm")

    def sniff(self, path: Path) -> bool:
        return path.suffix.lower() in {".xlsx", ".xlsm"}

    def load_raw(self, path: Path) -> Any:
        try:
            wb = load_workbook(filename=path, data_only=True)
            data = {}

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.values)

                if not rows:
                    data[sheet] = []
                    continue

                header = [str(h) for h in rows[0]]
                sheet_rows = []

                for row in rows[1:]:
                    sheet_rows.append({header[i]: row[i] for i in range(len(header))})

                data[sheet] = sheet_rows

            return data

        except Exception as exc:
            raise LoaderReadError(
                f"Failed to read XLSX file '{path}': {exc}"
            ) from exc

    def extract_metadata(self, raw_data: Any) -> Mapping[str, Any]:
        try:
            return {
                "format": self.FORMAT,
                "num_sheets": len(raw_data),
                "sheets": list(raw_data.keys()),
            }
        except Exception as exc:
            raise LoaderMetadataError(
                f"Failed to extract XLSX metadata: {exc}"
            ) from exc

    def to_universal(self, raw_data: Any, metadata: Mapping[str, Any]) -> Any:
        try:
            return raw_data
        except Exception as exc:
            raise LoaderNormalizationError(
                f"Failed to normalize XLSX data: {exc}"
            ) from exc
